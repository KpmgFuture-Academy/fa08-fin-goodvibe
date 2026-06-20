#!/usr/bin/env python
"""make_report.py — coach_eval 결과 CSV → 엑셀(.xlsx) 보고서.

3개 시트:
  1) 요약      — 모델별 정확도/false_OK/false_REJ/지연/비용 + 추천 + 한계
  2) 상세(피벗) — 사진별 정답 vs 모델별 예측 (정오답 색상 표시, false_OK 강조)
  3) 원본데이터 — 모델×사진 전체 행 (메시지/지연/토큰 포함)

사용:
    python tools/make_report.py --results test_photos/manifest_results.csv \
        --out test_photos/coach_model_report.xlsx
"""
from __future__ import annotations

import argparse
import csv
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# $/1M (input, output) — coach_eval 와 동일.
PRICES = {
    "gpt-4.1-mini": (0.40, 1.60),
    "gpt-4o-mini": (0.15, 0.60),
    "gemini-2.5-flash-lite": (0.10, 0.40),
    "gemini-2.5-flash": (0.30, 2.50),
    "gemini-3.1-flash-lite": (0.10, 0.40),
}

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
OK_FILL = PatternFill("solid", fgColor="C6EFCE")     # 정답 = 연두
BAD_FILL = PatternFill("solid", fgColor="FFEB9C")    # 오답(안전) = 노랑
FALSEOK_FILL = PatternFill("solid", fgColor="FF5B5B")  # false_OK = 빨강
FALSEOK_FONT = Font(color="FFFFFF", bold=True)
REC_FILL = PatternFill("solid", fgColor="C6EFCE")
TITLE_FONT = Font(bold=True, size=14)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _read(results_path: Path) -> list[dict]:
    with results_path.open(encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def _model_order(rows: list[dict]) -> list[str]:
    seen: list[str] = []
    for r in rows:
        if r["model"] not in seen:
            seen.append(r["model"])
    return seen


def _summary(rows: list[dict], model: str) -> dict:
    recs = [r for r in rows if r["model"] == model]
    valid = [r for r in recs if not r.get("error")]
    n = len(valid)
    errors = len(recs) - n
    if n == 0:
        return {"model": model, "n": 0, "errors": errors}
    exact = sum(1 for r in valid if r["status"] == r["expected"]) / n
    not_ok = [r for r in valid if r["expected"] != "ok"]
    false_ok = (sum(1 for r in not_ok if r["status"] == "ok") / len(not_ok)) if not_ok else 0.0
    should_ok = [r for r in valid if r["expected"] == "ok"]
    false_rej = (sum(1 for r in should_ok if r["status"] != "ok") / len(should_ok)) if should_ok else 0.0
    lat = [int(r["latency_ms"]) for r in valid if r.get("latency_ms")]
    avg_ms = sum(lat) / len(lat) if lat else 0
    price = PRICES.get(model)
    cost1k = None
    if price:
        pin, pout = price
        ptok = sum(int(r["prompt_tokens"] or 0) for r in valid) / n
        ctok = sum(int(r["completion_tokens"] or 0) for r in valid) / n
        cost1k = (ptok * pin + ctok * pout) / 1_000_000 * 1000
    return {"model": model, "n": n, "errors": errors, "exact": exact,
            "false_ok": false_ok, "false_rej": false_rej, "avg_ms": round(avg_ms), "cost1k": cost1k}


def _autofit(ws, widths: dict[int, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def build(results_path: Path, out_path: Path) -> None:
    rows = _read(results_path)
    models = _model_order(rows)
    summaries = [_summary(rows, m) for m in models]
    # 추천: false_OK 최소 → 정확도 최대 → 비용 최소
    ranked = [s for s in summaries if s["n"]]
    rec = min(ranked, key=lambda s: (s["false_ok"], -s["exact"], s.get("cost1k") or 9e9)) if ranked else None

    # 테스트셋 구성
    expected_counts = defaultdict(int)
    for r in rows:
        if r["model"] == models[0]:
            expected_counts[r["expected"]] += 1
    total_imgs = sum(expected_counts.values())

    wb = Workbook()

    # ── 시트 1: 요약 ──
    ws = wb.active
    ws.title = "요약"
    ws["A1"] = "사진 라이브 코칭 — 비전 모델 오판율 비교 보고서"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"생성일 {datetime.now():%Y-%m-%d %H:%M}  ·  테스트셋 {total_imgs}장 " \
               f"(ok {expected_counts.get('ok',0)} / wait {expected_counts.get('wait',0)} / adjust {expected_counts.get('adjust',0)})"
    ws["A2"].font = Font(color="808080", italic=True)

    hdr = ["모델", "N", "오류", "정확도", "false_OK", "false_REJ", "평균지연(ms)", "$/1k호출", "비고"]
    hr = 4
    for c, h in enumerate(hdr, 1):
        cell = ws.cell(hr, c, h)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT; cell.alignment = CENTER; cell.border = BORDER
    for i, s in enumerate(summaries):
        r = hr + 1 + i
        is_rec = rec is not None and s["model"] == rec["model"]
        vals = [
            s["model"], s.get("n", 0), s.get("errors", 0),
            s.get("exact"), s.get("false_ok"), s.get("false_rej"),
            s.get("avg_ms"), s.get("cost1k"),
            ("⭐ 추천" if is_rec else ("현재 운영" if s["model"] == "gpt-4.1-mini" else "")),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v); cell.border = BORDER; cell.alignment = CENTER
            if c == 4 or c == 6:  # 정확도, false_REJ
                cell.number_format = "0.0%"
            elif c == 5:  # false_OK
                cell.number_format = "0.0%"
                if isinstance(v, (int, float)):
                    if v > 0:
                        cell.fill = FALSEOK_FILL; cell.font = FALSEOK_FONT
                    else:
                        cell.fill = OK_FILL
            elif c == 8 and isinstance(v, (int, float)):  # 비용
                cell.number_format = '"$"0.00'
            if is_rec and c != 5:
                cell.fill = REC_FILL
        ws.cell(r, 1).alignment = Alignment(horizontal="left")
        ws.cell(r, 9).alignment = Alignment(horizontal="left")

    note_r = hr + len(summaries) + 2
    notes = [
        "지표 해설:",
        "  · false_OK  = 찍으면 안 되는 화면(물 찬 논·실내 등)을 'ok'로 통과시킨 비율. 부적합 증빙 통과 = 컴플라이언스 위험. ↓일수록 안전.",
        "  · false_REJ = 찍어도 되는 화면을 'ok' 안 준 비율. 어르신이 '왜 안 찍혀' 답답함. ↓일수록 UX 좋음.",
        "  · $/1k호출 = usage 토큰 실측 × 단가. 라이브 폴링 1000회 비용.",
        "",
        "결론 / 추천:",
        f"  · {rec['model'] if rec else '-'} 가 현재(gpt-4.1-mini) 대비 정확도↑·오판↓·비용 절반 이하 → 교체 권장.",
        "  · gemini-2.5-flash-lite 는 가장 싸지만 '물 찬 논'을 ok 통과(false_OK) — 미묘한 농촌 판정에서 약점.",
        "  · 주의: gemini-2.5-flash 는 thinking 기본 ON이라 max_tokens=80이면 빈 응답으로 실패. 채택 시 reasoning_effort='none' + 토큰 여유 필요.",
        "",
        "한계:",
        f"  · 표본 {total_imgs}장 + 대부분 AI 생성 이미지 → 방향성이지 통계적 확정 아님. 배포 후 실제 농가 사진 재검증 권장.",
    ]
    for i, t in enumerate(notes):
        cell = ws.cell(note_r + i, 1, t)
        if t.endswith(":"):
            cell.font = Font(bold=True)
    _autofit(ws, {1: 24, 2: 6, 3: 6, 4: 9, 5: 10, 6: 11, 7: 14, 8: 11, 9: 12})

    # ── 시트 2: 상세(피벗) ──
    ws2 = wb.create_sheet("상세_피벗")
    files: list[str] = []
    meta: dict[str, dict] = {}
    pred: dict[tuple[str, str], dict] = {}
    for r in rows:
        f = r["file"]
        if f not in meta:
            files.append(f)
            meta[f] = {"job_cd": r["job_cd"], "evidence_type": r["evidence_type"], "expected": r["expected"]}
        pred[(f, r["model"])] = r

    hdr2 = ["파일", "활동", "정답"] + [f"{m}\n예측" for m in models]
    for c, h in enumerate(hdr2, 1):
        cell = ws2.cell(1, c, h); cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); cell.border = BORDER
    for i, f in enumerate(files):
        r = 2 + i
        m = meta[f]
        act = (m["job_cd"] + "/" if m["job_cd"] else "") + m["evidence_type"]
        ws2.cell(r, 1, f).border = BORDER
        ws2.cell(r, 2, act).border = BORDER; ws2.cell(r, 2).alignment = CENTER
        exp = m["expected"]
        ec = ws2.cell(r, 3, exp); ec.border = BORDER; ec.alignment = CENTER; ec.font = Font(bold=True)
        for j, model in enumerate(models):
            pr = pred.get((f, model))
            cell = ws2.cell(r, 4 + j); cell.border = BORDER; cell.alignment = CENTER
            if not pr or pr.get("error"):
                cell.value = "ERR"; cell.fill = BAD_FILL; continue
            st = pr["status"]
            cell.value = st
            if st == exp:
                cell.fill = OK_FILL
            elif exp != "ok" and st == "ok":  # false OK
                cell.fill = FALSEOK_FILL; cell.font = FALSEOK_FONT
            else:
                cell.fill = BAD_FILL
    legend_r = len(files) + 3
    ws2.cell(legend_r, 1, "범례:").font = Font(bold=True)
    lg = [("연두 = 정답", OK_FILL), ("노랑 = 오답(안전: 통과 안 시킴)", BAD_FILL), ("빨강 = false_OK (부적합인데 통과 — 위험)", FALSEOK_FILL)]
    for i, (txt, fill) in enumerate(lg):
        c = ws2.cell(legend_r + 1 + i, 1, txt); c.fill = fill
        if fill is FALSEOK_FILL:
            c.font = FALSEOK_FONT
    _autofit(ws2, {1: 26, 2: 12, 3: 8, **{4 + j: 18 for j in range(len(models))}})
    ws2.freeze_panes = "D2"

    # ── 시트 3: 원본데이터 ──
    ws3 = wb.create_sheet("원본데이터")
    cols = ["model", "file", "job_cd", "evidence_type", "expected", "status",
            "can_capture", "message", "latency_ms", "prompt_tokens", "completion_tokens", "error"]
    for c, h in enumerate(cols, 1):
        cell = ws3.cell(1, c, h); cell.fill = HEADER_FILL; cell.font = HEADER_FONT; cell.border = BORDER
    for i, r in enumerate(rows):
        for c, k in enumerate(cols, 1):
            ws3.cell(2 + i, c, r.get(k, ""))
    _autofit(ws3, {1: 22, 2: 26, 3: 8, 4: 12, 5: 9, 6: 8, 7: 11, 8: 34, 9: 10, 10: 12, 11: 14, 12: 20})
    ws3.freeze_panes = "A2"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> int:
    for s in (sys.stdout, sys.stderr):
        try:
            s.reconfigure(encoding="utf-8")  # type: ignore[union-attr]
        except Exception:  # noqa: BLE001
            pass
    ap = argparse.ArgumentParser()
    ap.add_argument("--results", default="test_photos/manifest_results.csv")
    ap.add_argument("--out", default="test_photos/coach_model_report.xlsx")
    args = ap.parse_args()
    rp = Path(args.results)
    if not rp.exists():
        print(f"결과 CSV 없음: {rp}", file=sys.stderr)
        return 1
    build(rp, Path(args.out))
    print(f"엑셀 보고서 생성: {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
